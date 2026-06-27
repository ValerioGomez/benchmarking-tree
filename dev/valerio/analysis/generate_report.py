import json
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

def generate():
    json_path = "dev/valerio/analysis/benchmark_results_optimized.json"
    excel_path = "dev/valerio/analysis/metrics_results.xlsx"

    if not os.path.exists(json_path):
        print(f"[Report] Error: No se encontro el archivo de resultados: {json_path}")
        return

    with open(json_path, 'r') as f:
        data = json.load(f)

    # Preparar datos
    rows = []
    names_map = {
        "b_star_traditional": "Árbol B* (Tradicional)",
        "b_star_ml": "Árbol B* + ML (RMI Baseline - Antes)",
        "b_star_ml_optimized": "★ Árbol B* + ML (RMI Optimizado - Después)",
        "skip_list_traditional": "Skip List (Tradicional)",
        "skip_list_ml": "Skip List + ML (Baseline - Antes)",
        "skip_list_ml_optimized": "★ Skip List + ML (Optimizado - Después)"
    }

    for key, name in names_map.items():
        if key in data:
            item = data[key]
            rows.append({
                "Estructura": name,
                "Tiempo Const. (ms)": item["build_time_ms"],
                "Tiempo Total Búsqueda 10k (ms)": item["total_search_time_ms"],
                "Latencia Promedio (us)": item["avg_latency_us"],
                "Varianza (us^2)": item["variance_us"],
                "Desviación Estándar (us)": item["stddev_us"]
            })

    df = pd.DataFrame(rows)

    # Crear libro de excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Benchmarking"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    data_font = Font(name=font_family, size=11)
    bold_data_font = Font(name=font_family, size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Título del reporte
    ws["B2"] = "Estudio de Optimización ML: Learned Indexes ('Antes' vs 'Después')"
    ws["B2"].font = title_font
    ws.row_dimensions[2].height = 25

    # Escribir cabeceras
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=2): # Columna B es 2
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Escribir datos
    for row_idx, row in df.iterrows():
        current_row = row_idx + 5
        ws.row_dimensions[current_row].height = 20
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            # Formatear números
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = bold_data_font
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.0000'

    # Calcular ganancias
    ws.cell(row=13, column=2, value="Comparativa de Ganancia de Rendimiento (Latencia)").font = Font(name=font_family, size=12, bold=True, color="1F4E78")
    
    ws.cell(row=15, column=2, value="Arquitectura Comparada").font = header_font
    ws.cell(row=15, column=2).fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws.cell(row=15, column=2).border = thin_border
    
    ws.cell(row=15, column=3, value="Reducción de Latencia (%)").font = header_font
    ws.cell(row=15, column=3).fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    ws.cell(row=15, column=3).border = thin_border
    ws.cell(row=15, column=3).alignment = Alignment(horizontal="center")

    # B* ML Baseline vs Trad
    ws.cell(row=16, column=2, value="Árbol B* + ML (Base) vs Árbol B* Tradicional").font = data_font
    ws.cell(row=16, column=2).border = thin_border
    cell_b_star = ws.cell(row=16, column=3, value="=(E5-E6)/E5")
    cell_b_star.font = bold_data_font
    cell_b_star.border = thin_border
    cell_b_star.alignment = Alignment(horizontal="right")
    cell_b_star.number_format = '0.0%'

    # B* ML Optimizado vs Trad
    ws.cell(row=17, column=2, value="★ Árbol B* + ML (Optimizado) vs Árbol B* Trad.").font = data_font
    ws.cell(row=17, column=2).border = thin_border
    cell_b_star_opt = ws.cell(row=17, column=3, value="=(E5-E7)/E5")
    cell_b_star_opt.font = bold_data_font
    cell_b_star_opt.border = thin_border
    cell_b_star_opt.alignment = Alignment(horizontal="right")
    cell_b_star_opt.number_format = '0.0%'

    # B* ML Optimizado vs Baseline (Mejora Neta)
    ws.cell(row=18, column=2, value="★ Árbol B* + ML (Optimizado) vs RMI Baseline").font = data_font
    ws.cell(row=18, column=2).border = thin_border
    cell_b_star_gain = ws.cell(row=18, column=3, value="=(E6-E7)/E6")
    cell_b_star_gain.font = bold_data_font
    cell_b_star_gain.border = thin_border
    cell_b_star_gain.alignment = Alignment(horizontal="right")
    cell_b_star_gain.number_format = '0.0%'

    # Skip List ML Baseline vs Trad
    ws.cell(row=20, column=2, value="Skip List + ML (Base) vs Skip List Tradicional").font = data_font
    ws.cell(row=20, column=2).border = thin_border
    cell_skip = ws.cell(row=20, column=3, value="=(E8-E9)/E8")
    cell_skip.font = bold_data_font
    cell_skip.border = thin_border
    cell_skip.alignment = Alignment(horizontal="right")
    cell_skip.number_format = '0.0%'

    # Skip List ML Optimizada vs Trad
    ws.cell(row=21, column=2, value="★ Skip List + ML (Optimizada) vs Skip List Trad.").font = data_font
    ws.cell(row=21, column=2).border = thin_border
    cell_skip_opt = ws.cell(row=21, column=3, value="=(E8-E10)/E8")
    cell_skip_opt.font = bold_data_font
    cell_skip_opt.border = thin_border
    cell_skip_opt.alignment = Alignment(horizontal="right")
    cell_skip_opt.number_format = '0.0%'

    # Skip List ML Optimizada vs Baseline (Mejora Neta)
    ws.cell(row=22, column=2, value="★ Skip List + ML (Optimizada) vs ML Baseline").font = data_font
    ws.cell(row=22, column=2).border = thin_border
    cell_skip_gain = ws.cell(row=22, column=3, value="=(E9-E10)/E9")
    cell_skip_gain.font = bold_data_font
    cell_skip_gain.border = thin_border
    cell_skip_gain.alignment = Alignment(horizontal="right")
    cell_skip_gain.number_format = '0.0%'

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        if col[0].column == 1:
            ws.column_dimensions['A'].width = 3
            continue
        max_len = 0
        for cell in col:
            if cell.coordinate == "B2":
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Insertar Gráfico de Latencia de Búsqueda
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Latencia Promedio por Búsqueda (Microsegundos)"
    chart.y_axis.title = "Microsegundos (us)"
    chart.x_axis.title = "Estructura"

    # Latencia es la columna E (col 5). Hay 6 estructuras (filas 5 a 10)
    data_ref = Reference(ws, min_col=5, min_row=4, max_row=10)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=10)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "B25")

    # Guardar archivo
    wb.save(excel_path)
    print(f"[Report] Reporte Excel generado y formateado exitosamente en: {excel_path}")

if __name__ == "__main__":
    generate()
